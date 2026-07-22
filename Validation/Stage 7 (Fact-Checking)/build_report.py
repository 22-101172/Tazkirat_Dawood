"""Build the Stage 7 disease-area report workbook from the full-run results.

Self-contained: reads the full verdict CSV, applies the disease-area theming (a
heuristic keyword grouping of the book's archaic conditions into modern research
areas), writes the themed flat file, then builds the multi-sheet workbook. Rerun
after any new full run: `python build_report.py`.
"""
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "data/factcheck_retrieve_judge_v1_full.csv"
THEMED = "data/_full_themed.csv"
OUT = "data/Stage7_book_answers_by_disease.xlsx"

# Ordered longest-match-first; a condition is tagged by the first pattern it hits.
THEMES = [
    ("Cancer / tumors", r"tumor|swelling|indurat|\bmass|growth|cancer|neoplas|carcino|polyp"),
    ("Infection / antimicrobial / wounds", r"infect|ulcer|abscess|wound|\bpus|gangren|\bsore|sepsis|fistula|boil"),
    ("Digestive / liver / GI", r"diarrhea|colic|stomach|gastr|digest|dyspep|constipat|dysenter|worm|helminth|liver|hepat|spleen|splen|jaundice|bowel|intestin|nausea|vomit|flatulen|anorexia|hemorrhoid|appetite|abdom|belly|indigest"),
    ("Mental / neurological", r"melanchol|grief|anxiet|depress|mania|memory|amnesia|epilep|seizure|convuls|insomnia|sleep|anger|obsess|distress|nerve|neuro|paralys|hemipleg|palsy|cerebral|tremor|spasm|vertigo|dizz|faint|headache|migraine|apoplex"),
    ("Metabolic (obesity/diabetes)", r"obesit|weight|diabet|\bsugar|cholesterol|thirst|emaciat"),
    ("Respiratory", r"cough|asthma|phlegm|chest|lung|bronch|respir|dyspnea|pulmonar|pleuri|catarrh"),
    ("Cardiovascular / blood", r"heart|palpitat|pressure|hemorrhage|bleed|\bblood|cardiac|vascular|edema|dropsy|anemia"),
    ("Urinary / renal", r"kidney|urin|bladder|stone|calcul|gravel|renal|diuret|dysuria|nephr"),
    ("Pain / joints / bones", r"\bpain|gout|joint|arthr|sciatic|rheumat|\bback|bone|fracture|muscle|cramp|sprain"),
    ("Skin / hair", r"skin|eczema|leprosy|scab|itch|pruritus|dermat|rash|freckle|vitiligo|melasma|alopecia|\bhair|baldness|dandruff|wart|callus|burn|blister|acne"),
    ("Eye / ENT / oral", r"\beye|ophthalm|visual|vision|cataract|\bear\b|deaf|throat|tonsil|\bnose|nasal|gingiv|dental|\btooth|teeth|\bgum|oral|mouth"),
    ("Reproductive / urogenital", r"menstru|amenorrh|uter|womb|fertil|sperm|libido|aphrodisiac|sexual|labor|pregnan|placenta|fetus|erectile|impoten|menopaus|lactation|breast"),
    ("Inflammation / fever", r"inflammat|fever|febrile"),
    ("Antioxidant / detox / poisoning", r"toxin|toxic|poison|venom|antidote"),
]
AREAS = [t[0] for t in THEMES] + ["Other / general"]


def theme(condition):
    c = str(condition).lower()
    for name, pat in THEMES:
        if re.search(pat, c):
            return name
    return "Other / general"


d = pd.read_csv(SRC)
d["disease_area"] = d["disease_or_condition_en"].apply(theme)
d.to_csv(THEMED, index=False)

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUB_FONT = Font(name=FONT, italic=True, size=10, color="555555")
BASE_FONT = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- Read me
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
ws["A1"] = "Tadhkirat Dawood — what the book answers, by disease area"
ws["A1"].font = TITLE_FONT
lines = [
    "Full fact-check of all 7,362 herb-disease claims against modern literature (Europe PMC / PubMed).",
    "Model: gemini-3-flash-preview, retrieve-then-judge pipeline. Run completed 2026-07-22.",
    "",
    "This workbook reframes the results around Dr. Taher's question: which heavily-researched modern",
    "diseases does the book already answer? The 'Confirmed' sheet lists every claim modern evidence",
    "supports, grouped by disease area and ranked by how much research exists (db_total_hits).",
    "",
    "Verdict legend:",
    "  Strongly Agree / Agree — a modern study tested this plant on this condition and confirmed it.",
    "  Neutral — real condition with indirect/traditional support, but no direct modern trial (a lead).",
    "  Disagree / Strongly Disagree — evidence shows the plant ineffective or harmful for this use.",
    "  No Modern Equivalent — a purely humoral concept with no modern counterpart (untranslatable).",
    "  Insufficient Evidence — a real condition, but no study exists for this specific plant.",
    "",
    "Notes:",
    "  'db_total_hits' = number of papers Europe PMC holds for the plant+condition query = a proxy",
    "     for how heavily researched that pair is. 'papers_used' = papers passed to the judge.",
    "  'disease_area' is a heuristic keyword grouping of the book's archaic conditions (~2% unclassified).",
    "  Per-claim verdicts carry a ~7% run-to-run wobble on borderline cases; aggregate counts are stable.",
]
for i, t in enumerate(lines, start=3):
    ws[f"A{i}"] = t
    ws[f"A{i}"].font = SUB_FONT if t else BASE_FONT
ws.column_dimensions["A"].width = 100

# ---------------------------------------------------------------- All claims (data backbone)
cols = ["entry_id", "arabic_name", "scientific_name", "disease_area",
        "disease_or_condition_en", "treatment_method_en", "verdict", "confidence",
        "db_total_hits", "n_papers_after_filter", "needs_human_review"]
headers_all = ["entry_id", "arabic_name", "scientific_name", "disease_area", "condition",
               "treatment", "verdict", "confidence", "db_total_hits", "papers_used", "needs_review"]
wsa = wb.create_sheet("All claims")
wsa.append(headers_all)
dd = d[cols].sort_values(["disease_area", "db_total_hits"], ascending=[True, False])
for _, r in dd.iterrows():
    wsa.append([r[c] for c in cols])
VN, DN = "G", "D"  # verdict col, disease_area col in 'All claims'

# ---------------------------------------------------------------- By disease area (live COUNTIFS)
wsd = wb.create_sheet("By disease area")
wsd.sheet_view.showGridLines = False
wsd["A1"] = "Book claims by disease area"
wsd["A1"].font = TITLE_FONT
wsd["A2"] = "Counts computed from the 'All claims' sheet (full-run snapshot, 2026-07-22). Sorted by supported."
wsd["A2"].font = SUB_FONT
head = ["Disease area", "Total claims", "Supported", "  of which Strongly Agree",
        "Neutral (leads)", "Refuted", "Insufficient / No-equiv"]
hr = 4
for j, h in enumerate(head, start=1):
    c = wsd.cell(hr, j, h); c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = CENTER; c.border = BORDER

def counts(area):
    a = d[d.disease_area == area]
    sup_ = a.verdict.isin(["Agree", "Strongly Agree"]).sum()
    return dict(
        total=len(a), supported=int(sup_),
        sa=int((a.verdict == "Strongly Agree").sum()),
        neutral=int((a.verdict == "Neutral").sum()),
        refuted=int(a.verdict.isin(["Disagree", "Strongly Disagree"]).sum()),
        insuff=int(a.verdict.isin(["Insufficient Evidence", "No Modern Equivalent"]).sum()))

rows_sorted = sorted(AREAS, key=lambda a: counts(a)["supported"], reverse=True)
r = hr + 1
tot = dict(total=0, supported=0, sa=0, neutral=0, refuted=0, insuff=0)
for area in rows_sorted:
    c = counts(area)
    for k in tot: tot[k] += c[k]
    vals = [area, c["total"], c["supported"], c["sa"], c["neutral"], c["refuted"], c["insuff"]]
    for j, v in enumerate(vals, start=1):
        cell = wsd.cell(r, j, v); cell.border = BORDER
        cell.font = BASE_FONT
        cell.alignment = CENTER if j > 1 else Alignment(vertical="center")
    r += 1
# totals row (values)
tvals = ["TOTAL", tot["total"], tot["supported"], tot["sa"], tot["neutral"], tot["refuted"], tot["insuff"]]
for j, v in enumerate(tvals, start=1):
    c = wsd.cell(r, j, v); c.font = BOLD; c.border = BORDER
    c.alignment = CENTER if j > 1 else Alignment(vertical="center")
widths = [34, 12, 11, 24, 15, 10, 22]
for j, w in enumerate(widths, start=1):
    wsd.column_dimensions[get_column_letter(j)].width = w
wsd.freeze_panes = "A5"

# ---------------------------------------------------------------- Confirmed (Agree + Strongly Agree)
def write_claims(ws, df):
    hdr = ["entry_id", "arabic_name", "scientific_name", "disease_area", "condition",
           "treatment", "verdict", "confidence", "db_total_hits", "papers_used",
           "evidence_papers", "reason (Arabic)"]
    src = ["entry_id", "arabic_name", "scientific_name", "disease_area",
           "disease_or_condition_en", "treatment_method_en", "verdict", "confidence",
           "db_total_hits", "n_papers_after_filter", "evidence_papers", "reason"]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(1, j, h); c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = CENTER; c.border = BORDER
    for _, row in df.iterrows():
        ws.append([row.get(c) for c in src])
    ws.freeze_panes = "A2"
    w = [8, 16, 22, 26, 24, 24, 15, 10, 12, 10, 34, 60]
    for j, ww in enumerate(w, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ww
    for rr in range(2, ws.max_row + 1):
        for j in range(1, 13):
            cell = ws.cell(rr, j); cell.font = BASE_FONT
            cell.alignment = LEFT if j in (11, 12, 5, 6) else Alignment(vertical="top")

sup = d[d.verdict.isin(["Agree", "Strongly Agree"])].sort_values(
    ["disease_area", "db_total_hits"], ascending=[True, False])
write_claims(wb.create_sheet("Confirmed (Agree+SA)"), sup)

ref = d[d.verdict.isin(["Disagree", "Strongly Disagree"])].sort_values(
    ["disease_area", "db_total_hits"], ascending=[True, False])
write_claims(wb.create_sheet("Refuted (safety review)"), ref)

wb.save(OUT)
print("wrote", OUT, "| confirmed:", len(sup), "| refuted:", len(ref), "| all:", len(d))
